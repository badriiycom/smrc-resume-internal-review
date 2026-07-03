#!/usr/bin/env python3
"""SMRC KP Resume Screener.

Reads every resume in a folder, sends each to Claude, scores it against all
eight SMRC Key Personnel checklists in one pass, and writes a branded Excel
tracker. See RUNBOOK.md for setup and usage.

Two modes:
  --mode live  (default) One API call per resume, sequentially. Simple,
               results land as they complete.
  --mode batch Submits every resume as a single Anthropic Message Batch
               (~50% cheaper than live calls). Anthropic's SLA is "within
               24 hours"; batches this size typically finish much sooner.
               Safe to interrupt (Ctrl+C) and re-run the same command later
               — it remembers the batch ID and just checks status instead
               of resubmitting or re-billing.
"""

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import tempfile
import time
from pathlib import Path

from anthropic import Anthropic
from docx import Document as DocxDocument
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

from kp_criteria import KP_CRITERIA, OCI_WATCHLIST

DEFAULT_MODEL = "claude-sonnet-4-6"
MAX_RESUME_CHARS = 24000
SUPPORTED_EXTS = {".pdf", ".docx", ".txt", ".doc", ".rtf"}
DEFAULT_POLL_SECONDS = 30

CHECKPOINT_FIELDS = [
    "file",
    "candidate_name",
    "current_title_employer",
    "primary_kp_category",
    "meets_minimums",
    "fit_strength",
    "rationale",
    "secondary_categories",
    "multi_role_flag",
    "oci_flag",
    "oci_reason",
    "sanction_check_pending",
    "error",
]


# --------------------------------------------------------------------------
# Resume discovery & text extraction (shared by both modes)
# --------------------------------------------------------------------------


def find_resumes(folder):
    root = Path(folder)
    if not root.is_dir():
        sys.exit(f"Input folder not found: {folder}")
    return sorted(
        p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in SUPPORTED_EXTS
    )


def extract_text(path):
    ext = path.suffix.lower()
    if ext == ".pdf":
        reader = PdfReader(str(path))
        text = "\n".join((page.extract_text() or "") for page in reader.pages)
    elif ext == ".docx":
        doc = DocxDocument(str(path))
        text = "\n".join(p.text for p in doc.paragraphs)
    elif ext == ".txt":
        text = path.read_text(errors="ignore")
    elif ext in (".doc", ".rtf"):
        text = _convert_with_libreoffice(path)
    else:
        text = ""
    return text.strip()[:MAX_RESUME_CHARS]


def _convert_with_libreoffice(path):
    with tempfile.TemporaryDirectory() as tmp:
        subprocess.run(
            ["soffice", "--headless", "--convert-to", "txt:Text", "--outdir", tmp, str(path)],
            check=True,
            capture_output=True,
            timeout=60,
        )
        out = Path(tmp) / (path.stem + ".txt")
        return out.read_text(errors="ignore") if out.exists() else ""


def build_prompt(resume_text, filename):
    role_blocks = []
    for code, role in KP_CRITERIA.items():
        knockouts = "\n".join(f"  - {item}" for item in role["knockouts"])
        diffs = "\n".join(f"  - {item}" for item in role["differentiators"])
        role_blocks.append(
            f"### {code} — {role['title']}\n"
            f"Minimum requirements (ALL must be clearly, specifically evidenced):\n{knockouts}\n"
            f"Strategic differentiators (rate Strong/Some/None if evidenced):\n{diffs}"
        )
    roles_text = "\n\n".join(role_blocks)
    oci_text = "\n".join(f"  - {item}" for item in OCI_WATCHLIST)
    role_codes = ", ".join(f'"{c}"' for c in KP_CRITERIA)

    return f"""You are screening a resume for the CMS Supplemental Medical Review Contractor
(SMRC) recompete against Commence's eight Key Personnel (KP) role checklists below.

RESUME FILENAME: {filename}
RESUME TEXT:
\"\"\"
{resume_text}
\"\"\"

ROLE CHECKLISTS:
{roles_text}

OCI WATCHLIST (flag if the candidate currently appears to work for any of these, or for any
other apparent CMS medical-review competitor performing UPIC, Recovery Audit Contractor, MAC,
or QIO functions):
{oci_text}

Pick the single best-fit role as primary_kp_category — the role whose minimum requirements the
candidate comes closest to meeting, or where they most clearly qualify. If the resume does not
fit any of the eight roles at all, use "None".

Be conservative: only credit a requirement when the resume shows specific evidence (a license, a
degree, stated years of experience, a named program, a certification). Missing or ambiguous
evidence means the requirement is NOT met — do not infer or guess. Sanction/exclusion status
(OIG LEIE + SAM.gov) cannot be verified from a resume alone; always set sanction_check_pending to
true and let a human run that check separately — never claim it as verified.

Respond with ONLY a single JSON object (no markdown fencing, no commentary) with exactly these
keys:
{{
  "candidate_name": "string",
  "current_title_employer": "string",
  "primary_kp_category": one of [{role_codes}, "None"],
  "meets_minimums": "Y" | "N" | "Unclear",
  "fit_strength": "Strong" | "Moderate" | "Weak" | "None",
  "rationale": "string citing specific resume evidence for the primary category decision",
  "secondary_categories": ["array of other role codes this candidate could plausibly fill"],
  "oci_flag": "Y" | "N",
  "oci_reason": "string, empty if oci_flag is N",
  "sanction_check_pending": true
}}
"""


# --------------------------------------------------------------------------
# Model response parsing -> checkpoint row (shared by both modes)
# --------------------------------------------------------------------------


def parse_json_response(raw_text):
    match = re.search(r"\{.*\}", raw_text, re.DOTALL)
    if not match:
        raise ValueError(f"model did not return JSON: {raw_text[:200]!r}")
    return json.loads(match.group(0))


def row_from_result(file_key, result):
    row = {field: "" for field in CHECKPOINT_FIELDS}
    row["file"] = file_key
    row["candidate_name"] = result.get("candidate_name", "")
    row["current_title_employer"] = result.get("current_title_employer", "")
    row["primary_kp_category"] = result.get("primary_kp_category", "")
    row["meets_minimums"] = result.get("meets_minimums", "")
    row["fit_strength"] = result.get("fit_strength", "")
    row["rationale"] = result.get("rationale", "")
    secondary = result.get("secondary_categories") or []
    row["secondary_categories"] = ", ".join(secondary)
    row["multi_role_flag"] = "Y" if secondary else "N"
    row["oci_flag"] = result.get("oci_flag", "N")
    row["oci_reason"] = result.get("oci_reason", "")
    row["sanction_check_pending"] = "Y"
    return row


def row_from_error(file_key, message):
    row = {field: "" for field in CHECKPOINT_FIELDS}
    row["file"] = file_key
    row["error"] = str(message)
    return row


def message_text(message):
    return "".join(block.text for block in message.content if block.type == "text")


# --------------------------------------------------------------------------
# Checkpoint / Excel output (shared by both modes)
# --------------------------------------------------------------------------


def load_checkpoint(checkpoint_path):
    done = {}
    path = Path(checkpoint_path)
    if not path.exists():
        return done
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            done[row["file"]] = row
    return done


def append_checkpoint_rows(checkpoint_path, rows):
    checkpoint_path = Path(checkpoint_path)
    is_new = not checkpoint_path.exists()
    with open(checkpoint_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CHECKPOINT_FIELDS)
        if is_new:
            writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_excel(rows, out_path):
    rows = list(rows)
    wb = Workbook()

    tracker = wb.active
    tracker.title = "Candidate Tracker"
    headers = [
        "Candidate Name",
        "Current Title / Employer",
        "Primary KP Category",
        "Meets Minimums",
        "Fit Strength",
        "Rationale",
        "Secondary Categories",
        "Multi-Role Flag",
        "OCI Flag",
        "OCI Reason",
        "Sanction Check Pending",
        "Source File",
        "Error",
    ]
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(headers, start=1):
        cell = tracker.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r, row in enumerate(rows, start=2):
        values = [
            row.get("candidate_name", ""),
            row.get("current_title_employer", ""),
            row.get("primary_kp_category", ""),
            row.get("meets_minimums", ""),
            row.get("fit_strength", ""),
            row.get("rationale", ""),
            row.get("secondary_categories", ""),
            row.get("multi_role_flag", ""),
            row.get("oci_flag", ""),
            row.get("oci_reason", ""),
            row.get("sanction_check_pending", ""),
            row.get("file", ""),
            row.get("error", ""),
        ]
        for c, value in enumerate(values, start=1):
            tracker.cell(row=r, column=c, value=value).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    widths = [22, 28, 16, 14, 12, 60, 20, 14, 10, 30, 20, 50, 30]
    for col, width in enumerate(widths, start=1):
        tracker.column_dimensions[get_column_letter(col)].width = width
    tracker.freeze_panes = "A2"

    summary = wb.create_sheet("Category Summary")
    summary.cell(row=1, column=1, value="KP Category").font = Font(bold=True)
    summary.cell(row=1, column=2, value="Total Screened").font = Font(bold=True)
    summary.cell(row=1, column=3, value="Meets Minimums (Y)").font = Font(bold=True)
    summary.cell(row=1, column=4, value="Strong Fit").font = Font(bold=True)
    summary.cell(row=1, column=5, value="OCI Flags").font = Font(bold=True)

    last_row = len(rows) + 1
    categories = list(KP_CRITERIA) + ["None"]
    for i, code in enumerate(categories, start=2):
        summary.cell(row=i, column=1, value=code)
        summary.cell(
            row=i, column=2,
            value=f'=COUNTIF(\'Candidate Tracker\'!C2:C{last_row},A{i})',
        )
        summary.cell(
            row=i, column=3,
            value=(
                f'=COUNTIFS(\'Candidate Tracker\'!C2:C{last_row},A{i},'
                f'\'Candidate Tracker\'!D2:D{last_row},"Y")'
            ),
        )
        summary.cell(
            row=i, column=4,
            value=(
                f'=COUNTIFS(\'Candidate Tracker\'!C2:C{last_row},A{i},'
                f'\'Candidate Tracker\'!E2:E{last_row},"Strong")'
            ),
        )
        summary.cell(
            row=i, column=5,
            value=(
                f'=COUNTIFS(\'Candidate Tracker\'!C2:C{last_row},A{i},'
                f'\'Candidate Tracker\'!I2:I{last_row},"Y")'
            ),
        )
    for col, width in enumerate([16, 16, 20, 12, 12], start=1):
        summary.column_dimensions[get_column_letter(col)].width = width

    wb.save(out_path)


# --------------------------------------------------------------------------
# Live mode: one API call per resume, sequentially
# --------------------------------------------------------------------------


def screen_resume_live(client, model, path):
    file_key = str(path.resolve())
    try:
        text = extract_text(path)
        if not text:
            raise ValueError("no extractable text (empty or unsupported/scanned file)")
        response = client.messages.create(
            model=model,
            max_tokens=1024,
            messages=[{"role": "user", "content": build_prompt(text, path.name)}],
        )
        result = parse_json_response(message_text(response))
        return row_from_result(file_key, result)
    except Exception as exc:  # noqa: BLE001 - any failure must not kill the run
        return row_from_error(file_key, exc)


def run_live_mode(client, args, resumes):
    done = load_checkpoint(args.checkpoint)
    total = len(resumes)
    print(f"Screening {total} resume(s) (live mode) from {args.input}")

    checkpoint_path = Path(args.checkpoint)
    is_new = not checkpoint_path.exists()
    with open(checkpoint_path, "a", newline="", encoding="utf-8") as ckpt_f:
        writer = csv.DictWriter(ckpt_f, fieldnames=CHECKPOINT_FIELDS)
        if is_new:
            writer.writeheader()

        for i, path in enumerate(resumes, start=1):
            key = str(path.resolve())
            if key in done:
                print(f"[{i}/{total}] {path.name} — skipped (already in checkpoint)")
                continue
            print(f"[{i}/{total}] {path.name}")
            row = screen_resume_live(client, args.model, path)
            writer.writerow(row)
            ckpt_f.flush()
            done[key] = row
            if args.sleep:
                time.sleep(args.sleep)

    write_excel(done.values(), args.out)
    print(f"Done. Wrote {args.out} ({len(done)} candidates).")


# --------------------------------------------------------------------------
# Batch mode: submit everything as one Anthropic Message Batch
# --------------------------------------------------------------------------


def load_batch_state(state_path):
    path = Path(state_path)
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_batch_state(state_path, state):
    with open(state_path, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)


def build_batch_requests(resumes, model, done):
    """Extract text and build one batch request per not-yet-done resume.

    Extraction happens locally, before submission — a resume that fails to
    extract never reaches the API and is recorded as an error immediately.
    """
    requests = []
    id_to_file = {}
    error_rows = []
    for path in resumes:
        file_key = str(path.resolve())
        if file_key in done:
            continue
        try:
            text = extract_text(path)
            if not text:
                raise ValueError("no extractable text (empty or unsupported/scanned file)")
        except Exception as exc:  # noqa: BLE001
            error_rows.append(row_from_error(file_key, exc))
            continue
        custom_id = f"resume-{len(id_to_file):05d}"
        id_to_file[custom_id] = file_key
        requests.append(
            {
                "custom_id": custom_id,
                "params": {
                    "model": model,
                    "max_tokens": 1024,
                    "messages": [{"role": "user", "content": build_prompt(text, path.name)}],
                },
            }
        )
    return requests, id_to_file, error_rows


def run_batch_mode(client, args, resumes):
    state = load_batch_state(args.batch_state)

    if state is None:
        done = load_checkpoint(args.checkpoint)
        requests, id_to_file, error_rows = build_batch_requests(resumes, args.model, done)
        if error_rows:
            append_checkpoint_rows(args.checkpoint, error_rows)
            print(f"{len(error_rows)} resume(s) failed local text extraction — logged, not billed.")

        if not requests:
            print("Nothing new to submit — all resumes already in the checkpoint.")
            write_excel(load_checkpoint(args.checkpoint).values(), args.out)
            print(f"Wrote {args.out}.")
            return

        print(f"Submitting {len(requests)} resume(s) as one batch...")
        batch = client.messages.batches.create(requests=requests)
        state = {
            "batch_id": batch.id,
            "id_to_file": id_to_file,
            "model": args.model,
            "collected": False,
        }
        save_batch_state(args.batch_state, state)
        print(f"Batch submitted: {batch.id}")
        print(
            f"Progress is saved to {args.batch_state}. If this process is interrupted, "
            f"re-run the exact same command — it will pick up this batch instead of "
            f"resubmitting."
        )

    if state.get("collected"):
        print("This batch was already collected. Nothing to do.")
        write_excel(load_checkpoint(args.checkpoint).values(), args.out)
        print(f"Wrote {args.out}.")
        return

    batch_id = state["batch_id"]
    id_to_file = state["id_to_file"]

    while True:
        batch = client.messages.batches.retrieve(batch_id)
        counts = batch.request_counts
        print(
            f"Batch {batch_id}: {batch.processing_status} — "
            f"succeeded={counts.succeeded} errored={counts.errored} "
            f"processing={counts.processing} canceled={counts.canceled} "
            f"expired={counts.expired}"
        )
        if batch.processing_status == "ended":
            break
        if not args.wait:
            print(
                f"Still processing. Re-run the same command later to check again "
                f"(polling every {args.poll_interval}s would happen automatically with --wait)."
            )
            return
        time.sleep(args.poll_interval)

    new_rows = []
    for item in client.messages.batches.results(batch_id):
        file_key = id_to_file.get(item.custom_id, item.custom_id)
        if item.result.type == "succeeded":
            try:
                result = parse_json_response(message_text(item.result.message))
                new_rows.append(row_from_result(file_key, result))
            except Exception as exc:  # noqa: BLE001
                new_rows.append(row_from_error(file_key, exc))
        else:
            detail = getattr(item.result, "error", None)
            new_rows.append(row_from_error(file_key, f"batch request {item.result.type}: {detail}"))

    append_checkpoint_rows(args.checkpoint, new_rows)
    state["collected"] = True
    save_batch_state(args.batch_state, state)

    all_done = load_checkpoint(args.checkpoint)
    write_excel(all_done.values(), args.out)
    print(f"Done. Collected {len(new_rows)} result(s). Wrote {args.out} ({len(all_done)} candidates).")


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help="Folder of local resume files")
    parser.add_argument("--out", default="SMRC_Resume_Triage_Tracker.xlsx")
    parser.add_argument("--checkpoint", default="screen_checkpoint.csv")
    parser.add_argument("--limit", type=int, default=None, help="Only screen the first N resumes")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument(
        "--label", default=None, help="Optional run label, e.g. --label full-run (informational only)"
    )

    parser.add_argument(
        "--mode",
        choices=["live", "batch"],
        default="live",
        help="live: one call per resume, sequentially. "
        "batch: submit all resumes as one Anthropic Message Batch (~50%% cheaper, "
        "async, completes within 24h per Anthropic's SLA — usually much sooner).",
    )

    # live mode
    parser.add_argument("--sleep", type=float, default=0.5, help="[live] Seconds between API calls")

    # batch mode
    parser.add_argument("--batch-state", default="batch_state.json", help="[batch] Where to remember the submitted batch ID")
    parser.add_argument("--poll-interval", type=float, default=DEFAULT_POLL_SECONDS, help="[batch] Seconds between status checks while waiting")
    parser.add_argument("--wait", dest="wait", action="store_true", default=True, help="[batch] Block and poll until the batch finishes (default)")
    parser.add_argument("--no-wait", dest="wait", action="store_false", help="[batch] Check status once and exit if not finished; re-run later to check again")

    args = parser.parse_args()

    if not os.environ.get("ANTHROPIC_API_KEY"):
        sys.exit("ANTHROPIC_API_KEY is not set. See RUNBOOK.md step 3.")

    client = Anthropic()

    resumes = find_resumes(args.input)
    if args.limit:
        resumes = resumes[: args.limit]
    if not resumes:
        sys.exit(f"No resumes found in {args.input} (looked for {sorted(SUPPORTED_EXTS)})")

    label = f" ({args.label})" if args.label else ""
    if label:
        print(f"Run label:{label}")

    if args.mode == "live":
        run_live_mode(client, args, resumes)
    else:
        run_batch_mode(client, args, resumes)


if __name__ == "__main__":
    main()
